"""
RSU Two-State (CA/IL) Income Allocation Script

Reads RSU vesting data and California stay durations from an Excel file,
then allocates taxable income proportionally based on days spent in each state
from the grant date to each vesting date.

Rules:
  - From grant date through 12/31/2024: all days count as California.
  - In 2025: days listed in the 2025CADuration sheet are California;
    all remaining days are Illinois.
"""

import argparse
from datetime import date, timedelta
from pathlib import Path

import openpyxl


def load_vesting_events(ws):
    events = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        transaction_date = row[0].date() if hasattr(row[0], "date") else row[0]
        award_date = row[1].date() if hasattr(row[1], "date") else row[1]
        taxable_income = float(row[6])
        events.append({
            "transaction_date": transaction_date,
            "award_date": award_date,
            "shares": int(row[4]),
            "fmv": float(row[5]),
            "taxable_income": taxable_income,
        })
    return events


def load_ca_periods_2025(ws):
    periods = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        start = row[0].date() if hasattr(row[0], "date") else row[0]
        end = row[1].date() if hasattr(row[1], "date") else row[1]
        periods.append((start, end))
    return periods


def overlap_days(period_start, period_end, range_start, range_end):
    """Count the number of overlapping days (inclusive on both ends)."""
    lo = max(period_start, range_start)
    hi = min(period_end, range_end)
    if lo > hi:
        return 0
    return (hi - lo).days + 1


def compute_ca_days(grant_date, vest_date, ca_periods_2025):
    """
    Count California days in [grant_date, vest_date] (inclusive).
    - grant_date through 12/31/2024: all CA
    - 1/1/2025 onward: only the explicit CA periods from the sheet
    """
    end_of_2024 = date(2024, 12, 31)
    start_of_2025 = date(2025, 1, 1)

    ca_days = 0
    if grant_date <= end_of_2024:
        ca_days += overlap_days(grant_date, end_of_2024, grant_date, vest_date)

    for p_start, p_end in ca_periods_2025:
        effective_start = max(p_start, start_of_2025)
        ca_days += overlap_days(effective_start, p_end, grant_date, vest_date)

    return ca_days


def main():
    parser = argparse.ArgumentParser(description="RSU two-state allocation (CA/IL)")
    parser.add_argument(
        "--file",
        type=str,
        default=r"C:\Users\kunw\OneDrive - KUNW-MSFT\2025Tax\RSUTwoStateAllocation.xlsx",
        help="Path to the RSUTwoStateAllocation.xlsx file",
    )
    args = parser.parse_args()

    wb = openpyxl.load_workbook(Path(args.file), data_only=True)
    events = load_vesting_events(wb["2025RSU"])
    ca_periods = load_ca_periods_2025(wb["2025CADuration"])
    wb.close()

    header = (
        f"{'Vest Date':>12s}  {'Grant Date':>12s}  {'Taxable Income':>15s}  "
        f"{'Total Days':>10s}  {'CA Days':>8s}  {'IL Days':>8s}  "
        f"{'CA %':>7s}  {'IL %':>7s}  {'CA Income':>12s}  {'IL Income':>12s}"
    )
    separator = "-" * len(header)

    print("\nRSU Two-State Allocation (CA / IL)")
    print(separator)
    print(header)
    print(separator)

    total_taxable = 0.0
    total_ca_income = 0.0
    total_il_income = 0.0

    for ev in events:
        grant = ev["award_date"]
        vest = ev["transaction_date"]
        total_days = (vest - grant).days + 1  # inclusive
        ca_days = compute_ca_days(grant, vest, ca_periods)
        il_days = total_days - ca_days

        ca_pct = ca_days / total_days
        il_pct = il_days / total_days
        ca_income = ev["taxable_income"] * ca_pct
        il_income = ev["taxable_income"] * il_pct

        total_taxable += ev["taxable_income"]
        total_ca_income += ca_income
        total_il_income += il_income

        print(
            f"{vest.strftime('%m/%d/%Y'):>12s}  {grant.strftime('%m/%d/%Y'):>12s}  "
            f"${ev['taxable_income']:>14,.2f}  {total_days:>10d}  {ca_days:>8d}  {il_days:>8d}  "
            f"{ca_pct:>6.1%}  {il_pct:>6.1%}  ${ca_income:>11,.2f}  ${il_income:>11,.2f}"
        )

    print(separator)
    print(
        f"{'TOTAL':>12s}  {'':>12s}  ${total_taxable:>14,.2f}  "
        f"{'':>10s}  {'':>8s}  {'':>8s}  "
        f"{'':>7s}  {'':>7s}  ${total_ca_income:>11,.2f}  ${total_il_income:>11,.2f}"
    )
    print()


if __name__ == "__main__":
    main()
